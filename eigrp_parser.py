#!/Library/Frameworks/Python.framework/Versions/3.12/bin/python3

import textfsm
import argparse
import sys
from os import path
from openpyxl import Workbook

def main():

	eigrp_topology_template_filename = './textfsm_templates/eigrp_topology.textfsm'
	parser = argparse.ArgumentParser(description='EIGRP Topology Parser')
	parser.add_argument(
		'-i',
		type=str,
		help='Input File',
	)

	parser.add_argument(
		'-o',
		type=str,
		help='Output File',
	)

	args = parser.parse_args()

	if not path.exists(args.i):
		print('Input File Not Found')
		sys.exit()

	print('Parsing EIGRP Topology')

	with open(eigrp_topology_template_filename) as raw_file:
		data_parser = textfsm.TextFSM(raw_file)

	report_header = data_parser.header

	with open(args.i) as raw_file:
		report_body = data_parser.ParseText(raw_file.read())
	
	print('Generating Excel Report')

	wb = Workbook()
	ws = wb.active
	ws.title = 'EIGRP Topology'

	for col_index, cell_value in enumerate(report_header, start=1):
		ws.cell(row=1, column=col_index).value = cell_value

	for row_index, row in enumerate(report_body, start=2):
		for col_index, cell_value in enumerate(row, start=1):
			if isinstance(cell_value, list):
				cell_value = ';'.join(cell_value)
			elif cell_value.isnumeric():
				cell_value = int(cell_value)
			ws.cell(row=row_index, column=col_index).value = cell_value

	wb.save(args.o)

	print('Program Completed!!')


if __name__ == '__main__':
	main()